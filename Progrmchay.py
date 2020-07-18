import datetime
import openpyxl

book = openpyxl.load_workbook("C:\\Users\\gilad\\Desktop\\WeeklySchedule20200123-20200129.xlsx")
activeSheet = book.active
Gate1 = ["אולם 1"]
Gate2 = ["אולם 2"]
Gate3 = ["אולם 3"]
Gate4 = ["אולם 4"]
Gate5 = ["אולם 5"]
Gate6 = ["אולם 6"]
Gate7 = ["אולם 7"]
Gate8 = ["אולם 8"]
Gate9 = ["אולם 9"]
Gate10 = ["אולם 10"]
Gate11 = ["אולם 11"]
Gate12 = ["אולם 12"]
Gate13 = ["אולם 13"]
Gate14 = ["אולם 14"]
Gate15 = ["אולם 15"]
Gate16 = ["אולם 16"]
Gate17 = ["אולם 17"]

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "1":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate1 :
                Gate1.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "2":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate2 :
                Gate2.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "3":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate3 :
                Gate3.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "4":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate4 :
                Gate4.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "5":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate5 :
                Gate5.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "6":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate6 :
                Gate6.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "7":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate7 :
                Gate7.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "8":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate8:
                Gate8.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "9":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate9 :
                Gate9.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "10":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate10 :
                Gate10.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "11":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate11 :
                Gate11.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "12":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate12 :
                Gate12.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "13":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate13:
                Gate13.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "14":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate14 :
                Gate14.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "15":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate15 :
                Gate15.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "16":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate16 :
                Gate16.append(activeSheet.cell(row=i+1 , column=j).value)

for i in range(1,activeSheet.max_row+1):
    if activeSheet.cell(column=1 , row=i).value == "17":
        for j in range(2,activeSheet.max_column+1):
            if activeSheet.cell(row=i+1 , column=j).value == None:
                pass
            elif activeSheet.cell(row=i+1 , column=j).value not in Gate17 :
                Gate17.append(activeSheet.cell(row=i+1 , column=j).value)


Gates = [Gate1 ,Gate2,Gate3,Gate4,Gate5,Gate6,Gate7,Gate8,Gate9,Gate10,Gate11,Gate12,Gate13 , Gate14,Gate15,Gate16, Gate17]
WriteFile= openpyxl.Workbook()
sheetWrite =WriteFile.active
ReadFile = openpyxl.load_workbook("C:\\Users\\gilad\\Desktop\\Demo.xlsx")
sheetread = ReadFile.active



i=0
for Gate in Gates:
    i+=1
    j=1
    for movie in Gate:

        c= sheetWrite.cell(row=i, column=j)
        c.value=str(movie)
        j += 1
        WriteFile.save("C:\\Users\\gilad\\Desktop\\Demo.xlsx")







